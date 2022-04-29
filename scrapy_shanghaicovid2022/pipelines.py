# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill

from scrapy_shanghaicovid2022.items import InfectedNumberItem, RiskLevelItem


class ScrapyShanghaicovid2022Pipeline:

    FILE_NAME = '/Users/lijjin/Documents/work/git_ws/scrapy_shanghaicovid2022/九里亭三区1.xlsx'
    START = date(2022, 4, 18)

    def open_spider(self, spider):
        self.workbook = openpyxl.open(self.FILE_NAME)

    def close_spider(self, spider):
        self.workbook.close()

    def process_item(self, item, spider):
        if not isinstance(item, InfectedNumberItem):
            return item

        print(item)

        sheet = self.workbook['松江感染人数']
        delta = item['date'] - self.START
        col = delta.days + 2
        cell = sheet.cell(row=1, column=col)
        cell.value = item['date'].strftime('%m/%d')
        cell = sheet.cell(row=2, column=col)
        cell.value = item['totalControl']
        cell = sheet.cell(row=3, column=col)
        cell.value = item['totalSocial']
        cell = sheet.cell(row=4, column=col)
        cell.value = item['totalConfirmed']
        cell = sheet.cell(row=5, column=col)
        cell.value = item['totalAsymp']
        cell = sheet.cell(row=6, column=col)
        cell.value = item['controlConfirmed']
        cell = sheet.cell(row=7, column=col)
        cell.value = item['controlAsymp']
        cell = sheet.cell(row=8, column=col)
        cell.value = item['socialConfirmed']
        cell = sheet.cell(row=9, column=col)
        cell.value = item['socialAsymp']
        cell = sheet.cell(row=10, column=col)
        cell.value = item['asympToConfimed']
        self.workbook.save(self.FILE_NAME)
        return item

class RiskLevelPipeline:
    FILE_NAME = '/Users/lijjin/Documents/work/git_ws/scrapy_shanghaicovid2022/九里亭三区1.xlsx'
    START = date(2022, 4, 11)
    RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    LEVEL_COLOR = {'封控区':RED, '管控区':ORANGE, '防范区':YELLOW}

    def open_spider(self, spider):
        self.workbook = openpyxl.open(self.FILE_NAME)

    def close_spider(self, spider):
        self.workbook.close()

    def process_item(self, item, spider):
        if not isinstance(item, RiskLevelItem):
            return item
        print(item)

        delta = item['date'] - self.START
        col = delta.days + 2
        print(col)

        addresses = {}
        addresses.update(item['addresses'])
        sheet = self.workbook['九里亭三区']
        date_cell = sheet.cell(row=1, column=col)
        date_cell.value = item['date'].strftime('%m/%d')

        row = 2
        addr_cell = sheet.cell(row=row, column=1)
        while addr_cell.value is not None:
            cell = sheet.cell(row=row, column=col)
            addr = addr_cell.value
            if addr in addresses:
                level = addresses.pop(addr)
                cell.value = level
                cell.fill = self.LEVEL_COLOR[level]
            else:
                left_cell = sheet.cell(row=row, column=col-1)
                print(addr_cell.value)
                print(left_cell.value)
                cell.value = left_cell.value
                cell.fill = self.LEVEL_COLOR[cell.value]
            row += 1
            addr_cell = sheet.cell(row=row, column=1)

        # the addr doesn't exist in excel before, need to insert it
        for key in addresses:
            addr_cell = sheet.cell(row=row, column=1)
            addr_cell.value = key
            cell = sheet.cell(row=row, column=col)
            cell.value = addresses[key]
            cell.fill = self.LEVEL_COLOR[cell.value]

        self.workbook.save(self.FILE_NAME)
        return item