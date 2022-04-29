# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

from datetime import date
import openpyxl
from openpyxl.styles import PatternFill

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    #print_hi('PyCharm')
    workbook = openpyxl.open('/Users/lijjin/Documents/work/git_ws/scrapy_shanghaicovid2022/九里亭三区1 copy.xlsx')
    sheet = workbook['九里亭三区']
    start = date(2022, 4, 11)
    today = date.today()
    delta = today - start
    print(delta.days)
    cell = sheet.cell(row=2, column=15)
    cell.value = '封控区'
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell = sheet.cell(row=100, column=1)
    print(cell.value)
    workbook.save('/Users/lijjin/Documents/work/git_ws/scrapy_shanghaicovid2022/九里亭三区1 copy.xlsx')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
